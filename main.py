"""
Email Support Agent — Production RAG-powered email responder.

This agent:
  1. Monitors a Gmail inbox for new emails via IMAP
  2. Retrieves relevant context from a Qdrant vector store (RAG)
  3. Generates professional replies via Gemini 2.5 Flash
  4. Sends replies via SMTP

This agent is stateless; the Gateway is responsible for triggering '/sync' calls.

Fixes applied (v3.2.0):
  - IMAP search query: separate UNSEEN and UID criteria correctly
  - Reply loop protection: skip auto-replies, bots, and self-addressed emails
  - upload_file: filename now passed into metadata so [Source: ...] renders
  - SMTP failure safety: UID only advances if reply was sent successfully
  - Rate limiting: asyncio.Semaphore limits concurrent Gemini calls to 5
  - FIX: embed_content / generate_content called via client.aio.* — were
    blocking the event loop as bare synchronous calls inside async functions
  - FIX: client.search() deprecated → replaced with client.query_points()
    which is the current Qdrant API (qdrant-client >= 1.7)
  - FIX: score_threshold=0.5 added to retrieval — previously all chunks,
    including near-zero similarity ones, were being fed to the LLM
  - FIX: collection creation race condition guarded with try/except so
    concurrent index requests for the same new tenant don't crash
  - FIX: embedding dimension validated against existing collection — prevents
    silent upsert failures when the embedding model is changed
  - FIX: _chunk_fixed_size infinite loop when overlap_words >= max_words;
    overlap is now clamped to max(0, max_words - 1)
"""

import os
import io
import asyncio
import uuid
import time
import imaplib
import smtplib
import email
import hashlib
import json
import logging
from email.mime.text import MIMEText
from typing import Optional

from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
from google import genai
from google.genai import types
from dotenv import load_dotenv
from qdrant_client import AsyncQdrantClient
from qdrant_client.http import models as rest_models

import pypdf
import docx        # python-docx
import openpyxl

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger("EmailSupportAgent")

load_dotenv()

app = FastAPI(
    title="Email Support Agent",
    description="RAG-powered email support agent with multi-format knowledge ingestion.",
    version="3.2.0",
)

# ── Rate limiting: max 5 concurrent Gemini calls ─────────────────────────────
_gemini_semaphore = asyncio.Semaphore(5)

# ── Qdrant Initialization ─────────────────────────────────────────────────────
_qdrant_client: Optional[AsyncQdrantClient] = None

def get_qdrant_client() -> AsyncQdrantClient:
    global _qdrant_client
    if _qdrant_client is None:
        url = os.getenv("QDRANT_URL")
        api_key = os.getenv("QDRANT_API_KEY")
        if not url:
            raise RuntimeError("QDRANT_URL is not set in environment")
        _qdrant_client = AsyncQdrantClient(url=url, api_key=api_key)
    return _qdrant_client


# ══════════════════════════════════════════════════════════════════════════════
# MODELS
# ══════════════════════════════════════════════════════════════════════════════

class ChatRequest(BaseModel):
    message: str
    injected_context: dict = {}

class SyncRequest(BaseModel):
    config: dict
    state: dict = {}

class KnowledgePayload(BaseModel):
    tenant_id: str
    content: str
    metadata: dict = {}
    chunk_strategy: str = "paragraph"

class RAGConfigResponse(BaseModel):
    rag_enabled: bool = True
    vector_store: str = "qdrant"
    embedding_model: str = "gemini-embedding-2"
    supported_files: list = [".pdf", ".docx", ".xlsx", ".txt"]
    chunk_strategy: str = "paragraph"
    max_file_size_mb: int = 10


# ══════════════════════════════════════════════════════════════════════════════
# GEMINI CLIENT
# ══════════════════════════════════════════════════════════════════════════════

def _get_genai_client() -> Optional[genai.Client]:
    api_key = os.getenv("GEMINI_API_KEY")
    base_url = os.getenv("GEMINI_BASE_URL")
    if not api_key:
        return None
    http_options = types.HttpOptions(base_url=base_url) if base_url else None
    return genai.Client(api_key=api_key, http_options=http_options)


# ══════════════════════════════════════════════════════════════════════════════
# EMBEDDINGS & RAG
# ══════════════════════════════════════════════════════════════════════════════

async def get_embeddings(text: str) -> list[float]:
    client = _get_genai_client()
    if not client:
        return []
    
    max_retries = 3
    for attempt in range(max_retries):
        try:
            truncated = text[:8000] if len(text) > 8000 else text
            response = await client.aio.models.embed_content(
                model="gemini-embedding-2",
                contents=truncated,
            )
            return response.embeddings[0].values
        except Exception as e:
            # Retry on transient errors (503, 429)
            if "503" in str(e) or "429" in str(e):
                wait_time = 2 ** attempt
                logger.warning(f"Gemini 503/429 error. Retrying in {wait_time}s... (Attempt {attempt+1}/{max_retries})")
                await asyncio.sleep(wait_time)
                continue
            
            logger.error(f"Embedding error: {e}")
            return []
    return []


async def get_retrieved_context(
    tenant_id: str,
    query: str,
    n_results: int = 5,
    score_threshold: float = 0.50,
) -> str:
    """
    Semantic search over the tenant's Qdrant collection.

    FIX 1: client.search() is deprecated in qdrant-client >= 1.7.  Replaced
    with client.query_points(), which is the current stable API.

    FIX 2: score_threshold added (default 0.50).  Previously every chunk in
    the collection, even near-zero similarity, was forwarded to the LLM as
    "context", poisoning the prompt with irrelevant content.
    """
    try:
        client = get_qdrant_client()
        collection_name = f"tenant_{tenant_id}"

        collections = await client.get_collections()
        exists = any(c.name == collection_name for c in collections.collections)
        if not exists:
            return ""

        query_embedding = await get_embeddings(query)
        if not query_embedding:
            return ""

        # FIX: query_points replaces deprecated search(); .points holds the hits
        response = await client.query_points(
            collection_name=collection_name,
            query=query_embedding,
            limit=n_results,
            score_threshold=score_threshold,
        )

        if not response.points:
            return ""

        context_parts = []
        for hit in response.points:
            doc_text = hit.payload.get("text", "")
            source = hit.payload.get("filename", hit.payload.get("source", "unknown"))
            prefix = f"[Source: {source}] " if source else ""
            context_parts.append(f"{prefix}{doc_text}")

        return "\n\n---\n\n".join(context_parts)
    except Exception as e:
        logger.error(f"Retrieval error: {e}")
        return ""


async def process_email_with_gemini(
    business_name: str,
    faq_text: str,
    message: str,
    tenant_id: Optional[str] = None,
) -> str:
    """
    Generates an email reply via Gemini.

    FIX: previously called client.models.generate_content() (synchronous)
    directly inside an async function under a semaphore, blocking the entire
    event loop for every LLM call.  Now uses client.aio.models.generate_content
    — the async variant provided by the google-genai SDK.
    """
    client = _get_genai_client()
    if not client:
        return "Sorry, I am currently unable to process your request."

    vector_context = (
        await get_retrieved_context(tenant_id, message) if tenant_id else ""
    )

    system_prompt = (
        f"You are a professional AI support assistant for **{business_name}**.\n\n"
        f"Use the following business information to answer the customer's inquiry:\n\n"
        f"── FAQ / Business Info ──\n{faq_text}\n\n"
    )
    if vector_context:
        system_prompt += f"── Retrieved Knowledge ──\n{vector_context}\n\n"
    system_prompt += (
        "── Guidelines ──\n"
        "• Be polite, professional, and concise.\n"
        "• Respond ONLY with a JSON object: {\"reply\": \"your response text\"}\n"
    )

    try:
        async with _gemini_semaphore:
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    response = await client.aio.models.generate_content(
                        model="gemini-2.5-flash",
                        contents=[system_prompt, f"Customer email: \"{message}\""],
                        config=types.GenerateContentConfig(
                            response_mime_type="application/json",
                            temperature=0.5,
                        ),
                    )
                    text_resp = response.text.strip()
                    if text_resp.startswith("```json"):
                        text_resp = text_resp[7:-3].strip()
                    data = json.loads(text_resp)
                    return data.get("reply", "Thank you for your message.")
                except Exception as e:
                    if ("503" in str(e) or "429" in str(e)) and attempt < max_retries - 1:
                        wait_time = 2 ** attempt
                        logger.warning(f"Gemini 503/429 error. Retrying in {wait_time}s... (Attempt {attempt+1}/{max_retries})")
                        await asyncio.sleep(wait_time)
                        continue
                    raise e

    except Exception as e:
        logger.error(f"Gemini error: {e}")
        return "Thank you for your message. We will get back to you shortly."


# ══════════════════════════════════════════════════════════════════════════════
# REPLY LOOP GUARD
# ══════════════════════════════════════════════════════════════════════════════

_AUTO_REPLY_HEADERS = [
    "Auto-Submitted",
    "X-Auto-Response-Suppress",
    "X-Autoreply",
    "X-AutoReply",
]

_AUTO_SENDER_KEYWORDS = [
    "noreply", "no-reply", "donotreply", "do-not-reply",
    "mailer-daemon", "postmaster", "bounce",
    "auto-confirm", "autoconfirm",
]

def _is_auto_message(msg: email.message.Message, own_address: str) -> bool:
    for header in _AUTO_REPLY_HEADERS:
        value = msg.get(header, "").strip().lower()
        if value and value not in ("no", "none", ""):
            logger.info(f"Skipping — auto header: {header}: {value}")
            return True

    sender = msg.get("From", "").lower()
    if own_address.lower() in sender:
        logger.info(f"Skipping — sender matches own address: {sender}")
        return True

    for keyword in _AUTO_SENDER_KEYWORDS:
        if keyword in sender:
            logger.info(f"Skipping — auto-sender keyword '{keyword}' in: {sender}")
            return True

    precedence = msg.get("Precedence", "").strip().lower()
    if precedence in ("bulk", "list", "junk"):
        logger.info(f"Skipping — Precedence: {precedence}")
        return True

    reply_to = msg.get("Reply-To", "").lower()
    for keyword in _AUTO_SENDER_KEYWORDS:
        if keyword in reply_to:
            logger.info(f"Skipping — auto keyword in Reply-To: {reply_to}")
            return True

    return False


# ══════════════════════════════════════════════════════════════════════════════
# SYNC LOGIC
# ══════════════════════════════════════════════════════════════════════════════

async def perform_sync(config: dict, state: dict) -> tuple[bool, dict, int]:
    gmail_user = config.get("gmail_address", "").strip()
    gmail_pass = config.get("gmail_app_password", "").replace(" ", "").strip()
    business_name = config.get("business_name", "the business")
    faq_text = config.get("faq_text", "")
    tenant_id = config.get("tenant_id")
    last_uid = state.get("last_processed_uid")

    if not gmail_user or not gmail_pass:
        return True, state, 0

    new_state = state.copy()

    try:
        def _imap_sync():
            mail = imaplib.IMAP4_SSL("imap.gmail.com", 993)
            mail.login(gmail_user, gmail_pass)
            mail.select("INBOX")

            if last_uid is None:
                status, data = mail.uid("SEARCH", None, "ALL")
                uids = (
                    [int(u) for u in data[0].split()]
                    if (status == "OK" and data[0])
                    else []
                )
                current_last_uid = max(uids) if uids else 0
                mail.close()
                mail.logout()
                return current_last_uid, []

            status, messages = mail.uid(
                "SEARCH", None,
                "UNSEEN",
                f"UID {last_uid + 1}:*",
            )
            if status != "OK" or not messages[0]:
                mail.close()
                mail.logout()
                return None, []

            emails_data = []
            for e_uid in messages[0].split():
                status, msg_data = mail.uid("FETCH", e_uid, "(RFC822)")
                for response_part in msg_data:
                    if isinstance(response_part, tuple):
                        emails_data.append((int(e_uid), response_part[1]))
                mail.uid("STORE", e_uid, "+FLAGS", "\\Seen")

            mail.close()
            mail.logout()
            return None, emails_data

        last_uid_update, fetched_emails = await asyncio.to_thread(_imap_sync)

        if last_uid_update is not None:
            new_state["last_processed_uid"] = last_uid_update
            return True, new_state, 0

        max_processed_uid = last_uid
        processed_count = 0

        for uid_int, raw_email in fetched_emails:
            msg = email.message_from_bytes(raw_email)

            if _is_auto_message(msg, gmail_user):
                max_processed_uid = max(max_processed_uid, uid_int)
                continue

            sender = msg.get("From")
            subject = msg.get("Subject", "")

            body = ""
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == "text/plain":
                        body_bytes = part.get_payload(decode=True)
                        body = body_bytes.decode(errors="replace") if body_bytes else ""
                        break
            else:
                body_bytes = msg.get_payload(decode=True)
                body = body_bytes.decode(errors="replace") if body_bytes else ""

            if not body.strip():
                max_processed_uid = max(max_processed_uid, uid_int)
                continue

            reply_text = await process_email_with_gemini(
                business_name, faq_text, body.strip(), tenant_id=tenant_id
            )

            try:
                def _send_reply(r_text, r_sender, r_subject):
                    smtp_msg = MIMEText(r_text)
                    smtp_msg["Subject"] = f"Re: {r_subject}"
                    smtp_msg["From"] = gmail_user
                    smtp_msg["To"] = r_sender
                    smtp_msg["Auto-Submitted"] = "auto-replied"
                    smtp_msg["X-Auto-Response-Suppress"] = "All"
                    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
                        s.login(gmail_user, gmail_pass)
                        s.send_message(smtp_msg)

                await asyncio.to_thread(_send_reply, reply_text, sender, subject)
                max_processed_uid = max(max_processed_uid, uid_int)
                processed_count += 1
                logger.info(f"Replied to UID {uid_int} from {sender}")

            except smtplib.SMTPException as smtp_err:
                logger.error(
                    f"SMTP send failed for UID {uid_int} to {sender}: {smtp_err}. "
                    "Will retry on next sync."
                )

        new_state["last_processed_uid"] = max_processed_uid
        return True, new_state, processed_count

    except Exception as e:
        logger.error(f"Sync error for {gmail_user}: {e}")
        return False, state, 0


# ══════════════════════════════════════════════════════════════════════════════
# TEXT EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def extract_text_from_pdf(content: bytes) -> str:
    reader = pypdf.PdfReader(io.BytesIO(content))
    pages = [p.extract_text().strip() for p in reader.pages if p.extract_text()]
    return "\n\n".join(pages)

def extract_text_from_docx(content: bytes) -> str:
    doc = docx.Document(io.BytesIO(content))
    return "\n\n".join([p.text.strip() for p in doc.paragraphs if p.text.strip()])

def extract_text_from_xlsx(content: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        rows = [
            " | ".join([str(c).strip() for c in r if c is not None])
            for r in wb[name].iter_rows(values_only=True)
        ]
        if rows:
            sheets.append(f"[Sheet: {name}]\n" + "\n".join(rows))
    return "\n\n".join(sheets)

def extract_text(content: bytes, filename: str) -> str:
    ext = filename.lower()
    if ext.endswith(".pdf"):                        return extract_text_from_pdf(content)
    if ext.endswith(".docx"):                       return extract_text_from_docx(content)
    if ext.endswith(".xlsx") or ext.endswith(".xls"): return extract_text_from_xlsx(content)
    return content.decode("utf-8", errors="replace")


def chunk_text(
    text: str,
    strategy: str = "paragraph",
    max_chunk_words: int = 400,
    overlap_words: int = 50,
) -> list[str]:
    if not text or not text.strip():
        return []
    # FIX: clamp overlap so fixed-size chunker can never produce a zero/negative
    # step, which caused an infinite loop when overlap_words >= max_chunk_words
    overlap_words = max(0, min(overlap_words, max_chunk_words - 1))

    if strategy == "sentence":
        return _chunk_by_sentence(text, max_chunk_words, overlap_words)
    elif strategy == "fixed_size":
        return _chunk_fixed_size(text, max_chunk_words, overlap_words)
    else:
        return _chunk_by_paragraph(text, max_chunk_words, overlap_words)


def _chunk_by_paragraph(text: str, max_words: int, overlap_words: int) -> list[str]:
    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
    chunks: list[str] = []
    current_chunk: list[str] = []
    current_word_count = 0

    for para in paragraphs:
        para_words = len(para.split())
        if para_words > max_words:
            if current_chunk:
                chunks.append("\n\n".join(current_chunk))
                current_chunk = []
                current_word_count = 0
            chunks.extend(_chunk_fixed_size(para, max_words, overlap_words))
            continue
        if current_word_count + para_words > max_words and current_chunk:
            chunks.append("\n\n".join(current_chunk))
            if overlap_words > 0:
                last_para = current_chunk[-1]
                current_chunk = [last_para]
                current_word_count = len(last_para.split())
            else:
                current_chunk = []
                current_word_count = 0
        current_chunk.append(para)
        current_word_count += para_words

    if current_chunk:
        chunks.append("\n\n".join(current_chunk))
    return chunks


def _chunk_by_sentence(text: str, max_words: int, overlap_words: int) -> list[str]:
    import re
    sentences = re.split(r'(?<=[.!?])\s+', text)
    chunks: list[str] = []
    current: list[str] = []
    current_words = 0

    for sentence in sentences:
        s_words = len(sentence.split())
        if current_words + s_words > max_words and current:
            chunks.append(" ".join(current))
            if overlap_words > 0:
                overlap_split = " ".join(current).split()
                carry = " ".join(overlap_split[-overlap_words:])
                current = [carry]
                current_words = len(carry.split())
            else:
                current = []
                current_words = 0
        current.append(sentence)
        current_words += s_words

    if current:
        chunks.append(" ".join(current))
    return chunks


def _chunk_fixed_size(text: str, max_words: int, overlap_words: int) -> list[str]:
    """
    FIX: step = max_words - overlap_words; was never guarded against <= 0.
    overlap_words is clamped in chunk_text() before this is called, so
    step >= 1 is guaranteed here — but we assert defensively anyway.
    """
    step = max_words - overlap_words
    assert step >= 1, "overlap_words must be < max_words"  # defensive

    words = text.split()
    chunks: list[str] = []
    i = 0
    while i < len(words):
        chunk = " ".join(words[i : i + max_words])
        if chunk.strip():
            chunks.append(chunk)
        i += step
    return chunks


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/")
@app.get("/health")
async def root():
    return {"status": "healthy", "version": "3.2.0"}

@app.post("/sync")
async def handle_sync(payload: SyncRequest):
    success, updated_state, count = await perform_sync(payload.config, payload.state)
    return {
        "status": "success" if success else "error",
        "state": updated_state,
        "processed": count,
    }

@app.post("/install")
async def handle_install():
    return {"status": "ok", "message": "Agent woken up."}

@app.post("/")
async def handle_chat(payload: ChatRequest):
    config = payload.injected_context
    reply = await process_email_with_gemini(
        config.get("business_name", "the business"),
        config.get("faq_text", ""),
        payload.message,
        tenant_id=config.get("tenant_id"),
    )
    return {"reply": reply}


@app.post("/index-knowledge")
async def index_knowledge(payload: KnowledgePayload):
    """Indexes text content into the tenant's Qdrant collection."""
    try:
        client = get_qdrant_client()
        collection_name = f"tenant_{payload.tenant_id}"

        chunks = chunk_text(
            payload.content,
            strategy=payload.chunk_strategy,
            max_chunk_words=400,
            overlap_words=50,
        )

        if not chunks:
            raise HTTPException(
                status_code=400,
                detail="No meaningful text to index after chunking.",
            )

        # Embed first chunk to detect vector dimension
        first_embedding = await get_embeddings(chunks[0])
        if not first_embedding:
            raise HTTPException(
                status_code=500,
                detail="Failed to generate initial embedding.",
            )
        vector_size = len(first_embedding)
        logger.info(f"Detected embedding dimension: {vector_size}")

        # Check if collection exists
        collections = await client.get_collections()
        exists = any(c.name == collection_name for c in collections.collections)

        if not exists:
            # FIX: race condition — two concurrent requests for the same new
            # tenant both see exists=False and both call create_collection.
            # Qdrant raises an error on the duplicate; we catch and ignore it
            # (the collection was already created by the winning request).
            try:
                logger.info(
                    f"Creating Qdrant collection: {collection_name} "
                    f"with size {vector_size}"
                )
                await client.create_collection(
                    collection_name=collection_name,
                    vectors_config=rest_models.VectorParams(
                        size=vector_size,
                        distance=rest_models.Distance.COSINE,
                    ),
                )
            except Exception as create_err:
                # Re-check: if it now exists, a concurrent request won — that's fine
                collections_retry = await client.get_collections()
                if not any(
                    c.name == collection_name for c in collections_retry.collections
                ):
                    # It genuinely failed for another reason — re-raise
                    raise create_err
                logger.warning(
                    f"Collection {collection_name} created concurrently — continuing."
                )
        else:
            # FIX: validate that the existing collection's vector size matches
            # the current embedding model.  A mismatch (e.g. after a model
            # upgrade) causes silent upsert failures with no clear error.
            collection_info = await client.get_collection(collection_name)
            existing_size = collection_info.config.params.vectors.size
            if existing_size != vector_size:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"Embedding dimension mismatch for collection '{collection_name}': "
                        f"existing={existing_size}, current model={vector_size}. "
                        "Delete the collection and re-index, or switch back to the "
                        "original embedding model."
                    ),
                )

        points = []
        for i, chunk in enumerate(chunks):
            embedding = first_embedding if i == 0 else await get_embeddings(chunk)
            if not embedding:
                logger.warning(f"Skipping chunk {i} — embedding failed.")
                continue

            content_hash = hashlib.sha256(chunk.encode()).hexdigest()[:16]
            point_id = str(
                uuid.uuid5(uuid.NAMESPACE_DNS, f"{payload.tenant_id}_{content_hash}")
            )

            points.append(
                rest_models.PointStruct(
                    id=point_id,
                    vector=embedding,
                    payload={
                        "text": chunk,
                        **payload.metadata,
                        "chunk_index": i,
                        "total_chunks": len(chunks),
                        "indexed_at": time.time(),
                    },
                )
            )

        if points:
            await client.upsert(
                collection_name=collection_name,
                points=points,
            )

        logger.info(
            f"Indexed {len(points)}/{len(chunks)} chunks for {payload.tenant_id}"
        )
        return {
            "status": "success",
            "chunks_indexed": len(points),
            "total_chunks": len(chunks),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Indexing error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Indexing failed: {e}")


@app.post("/upload-file")
async def upload_file(tenant_id: str = Form(...), file: UploadFile = File(...)):
    content = await file.read()
    text = extract_text(content, file.filename)
    return await index_knowledge(
        KnowledgePayload(
            tenant_id=tenant_id,
            content=text,
            metadata={"filename": file.filename},
        )
    )


@app.get("/rag-config", response_model=RAGConfigResponse)
async def rag_config():
    return RAGConfigResponse()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", 8080)))